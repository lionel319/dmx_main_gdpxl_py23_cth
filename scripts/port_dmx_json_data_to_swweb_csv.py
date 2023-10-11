#!/usr/bin/env python

## @package port_dmx_json_data_to_swweb_csv
## @brief Reads in all the json CLOSE data from dmx, and generates it in a csv format that is needed by sw-web/ice app to update the django database.
## @author Lionel Tan Yoke-Liang (yltan@altera.com)
## @version 1.1
## @date 30 July 2016
## @copyright Intel PSG


import os
import sys
import logging
from argparse import ArgumentParser
import json
from pprint import pprint
import csv 
import openpyxl

LOGGER = logging.getLogger(__name__)

## Brief Description For Function
##
## @param varname a @c string . description of the param
## @exception dmx::ecolib::family::FamilyError description of the exception
## @return return @c True if lalala, else @c False
def main():
    ''' '''
    parser = ArgumentParser()
    parser.add_argument('--debug', required=False, action='store_true', default=False, help="More verbose by printing out debugging messages.")
    parser.add_argument('-i','--it_file', required=True, default=False, help="ip_type.json file from dmx.")
    parser.add_argument('-d','--dbm_file', required=True, default=False, help="deliverables_by_milestone.json file from dmx.")

    args = parser.parse_args()

    if args.debug:
        logging.basicConfig(format='[%(asctime)s] %(levelname)s:%(message)s', level=logging.DEBUG)         
    else:
        logging.basicConfig(format='[%(asctime)s] %(levelname)s:%(message)s', level=logging.INFO)    


    #f1 = 'fm4_ip_types.json'
    #f2 = 'fm4_deliverables_by_milestone.json'

    ipt = {}
    dbm = {}

    if args.it_file:
        with open(f1) as f:
            ipt = json.load(f)

    if args.dbm_file:
        with open(f2) as f:
            dbm = json.load(f)
        add_milestone_99(dbm)

    #dbl = gen_dict_by_libtypes(ipt, dbm)

    gen_roadmap(dbm, 'roadmap.csv')
    gen_roadmap_by_vt(dbm, ipt, 'roadmap_by_vt.csv')


def gen_roadmap_by_vt(dbm, ipt, outfile):

    wb = openpyxl.workbook.Workbook()
    ws = wb.create_sheet(index=0)

    vts = sorted(ipt.keys())

    r = 0
    c = 0

    ### HEADER
    ws.cell(row=r, column=c).value = 'milestone'
    c += 1
    for vt in vts:
        ws.cell(row=r, column=c).value = vt
        c += 1
    
    ### Start populating milestone
    for tms in sorted(dbm.keys()):
        ms_start_row = ws.get_highest_row()
        c = 0
        ws.cell(row=ms_start_row, column=c).value = tms
        for i, vt in enumerate(vts):
            r = ms_start_row
            c = i + 1
            for libtype in ipt[vt]:
                if libtype in dbm[tms]:
                    ws.cell(row=r, column=c).value = libtype.upper()
                    r += 1

    tmpfile = './roadmap_by_vt.xlsx'
    wb.save(tmpfile)
    worksheet_to_csv(ws, outfile)



def add_milestone_99(dbm):
    all_libtypes = []
    for tms in dbm:
        thread, ms = tms.split("REL")
        for libtype in dbm[tms]:
            all_libtypes.append(libtype.upper())
    dbm[thread+'REL99'] = list(set(all_libtypes))


def gen_roadmap(dbm, outfile):

    desc = {
        '0.5' : 'Spec Complete',
        '1.0' : 'Baseline Architecture Complete',
        '1.5' : 'Multi Sector', 
        '2.0' : 'Iinitial Fullchip Assembly',
        '2.5' : 'Minichip feature complete Verification',
        '3.0' : 'Fullchip Integration Complete',
        '3.5' : 'Verification 80%',
        '4.0' : 'Frozen / TMA1 start',
        '4.5' : 'TMA2 Hand Over',
        '5.0' : 'Tapeout',
        '99'  : 'reference'
    }

    wb = openpyxl.workbook.Workbook()
    ws = wb.create_sheet(index=0)

    all_libtypes = []
    c = 0
    for tms in sorted(dbm.keys()):
        thread, ms = tms.split('REL')
        r = 0

        ### Header
        ws.cell(row=r, column=c).value = tms

        ### Desc
        r += 1
        ws.cell(row=r, column=c).value = desc[ms]

        ### Libtypes
        for libtype in sorted(dbm[tms]):
            r += 1
            ws.cell(row=r, column=c).value = libtype.upper()
            all_libtypes.append(libtype.upper())

        c += 1

    

    tmpfile = './roadmap.xlsx'
    wb.save(tmpfile)
    worksheet_to_csv(ws, outfile)


def incrow(row):
    return str(int(row) + 1)

def inccol(col):
    return chr(ord(col) + 1)


def gen_dict_by_libtypes(ipt, dbm):
    '''
    ipt = {
        "asic" : ["IPSPEC", "CDL", ...]
        "custom" : [... ... ...],
        ... ... ...
    } # ip type

    dbm = {
        "FM4revAREL1.0" : ["IPSPEC", "CDL"],
        "FM4revAREL2.0" : [... ... ...],
        ... ... ...
    } # deliverables by milestone

    ret = {
        "IPSPEC" : { 
            "vt" : ["asic", "custom", ...],
            "ms" : ["FM4revAREL1.0", "FM4revAREL2.0", ...]
        },
        "CDL" : {
            "vt" : ["fc", "custom", ...],
            "ms" : ["FM4revAREL3.0", "FM4revAREL4.0", ...]
        },
        ... ... ..
    } # ret

    '''
    ret = {}
    for vt in ipt:
        for libtype in ipt[vt]:
            ret.setdefault(libtype, {"vt":[], "ms":[]})
            if vt not in ret[libtype]['vt']:
                ret[libtype]['vt'].append(vt)

    for tms in dbm:
        [thread, ms] = tms.split('REL')
        for libtype in dbm[tms]:
            if libtype not in ret:
                LOGGER.error("ERROR: libtype {} found in {}, but not in ip_type.json!!!".format(libtype, tms))
            else:
                ret[libtype]['ms'].append(tms)

    return ret


def worksheet_to_csv(ws, outfile):
    with open(outfile, 'wb') as f:
        c = csv.writer(f)
        for r in ws.rows:
            c.writerow([cell.value for cell in r])
    LOGGER.info("CSV file {} generated ...".format(outfile))

if __name__ == '__main__':
    main()


